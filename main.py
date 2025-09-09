import re
import streamlit as st
import pandas as pd
from typing import Optional
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ====== 标准库 ======
MYLAR_CATS = ["Mylar", "MB Mylar", "DB Mylar", "KB Mylar", "Touchpad Mylar", "D Cover Mylar", "TCON Mylar"]
COLORS = ["Black", "Yellow", "Blue"]
MATERIALS = ["PC", "PET", "PVC", "Acrylic"]
FP_FLAGS = ["W/ FP", "W/O FP"]
ADHESIVES = ["3M9495", "3M300LSE", "3M9448A", "3M9448B", "3M200MP", "DSTT-13N", "DSTT-7N", "SDK7100"]
OTHERS = ["Object", "Gluing"]

# 只清洗这些列（与你截图一致）
TARGET_COLS = ["物料简称", "名称", "项目名称", "颜色", "材质", "长L(mm)", "宽W(mm)", "厚H(mm)", "是否带指纹", "背胶型号",
               "其它特殊属性"]


# ====== 小工具 ======
def smart_title(text: str) -> str:
    """首字母大写，保留缩写大写"""
    specials = {"MB", "DB", "KB", "TCON", "PC", "PET", "PVC"}
    text = text.replace("-", " ").replace("_", " ")
    words = text.split()
    return " ".join([w.upper() if w.upper() in specials else w.capitalize() for w in words])


def fuzzy_one(s: str, pool: list, threshold: int) -> Optional[str]:
    if not s:
        return None
    match = process.extractOne(s, pool, scorer=fuzz.ratio)
    if match and match[1] >= threshold:
        return match[0]
    return None


# ====== 列专用清洗器（只对指定列启用对应规则，避免误清洗）======
# 1) 物料简述 / 项目名称：只做 Mylar 类别归一；不做颜色/材质映射，避免人名/项目名被误改
def clean_mylar_name(val: str):
    if pd.isna(val) or not str(val).strip():
        return val, False, "none"
    raw = str(val).strip()
    txt_u = raw.upper()

    # 中文/混合关键词（仅限 Mylar 类别）
    if "触摸板" in raw: return "Touchpad Mylar", True, "cn_keyword"
    if "键盘" in raw: return "KB Mylar", True, "cn_keyword"
    if "主板" in raw: return "MB Mylar", True, "cn_keyword"
    if "副板" in raw or "DDR" in txt_u: return "DB Mylar", True, "cn_keyword"
    if "麦拉" in raw or "MYLAR" in txt_u:  # 泛指麦拉
        # 若包含 MB/DB/KB/TP 等再归类，否则归为 Mylar
        if " MB" in f" {txt_u} " or txt_u.startswith("MB "):
            return "MB Mylar", True, "keyword_rule"
        if " DB" in f" {txt_u} ":
            return "DB Mylar", True, "keyword_rule"
        if " KB" in f" {txt_u} ":
            return "KB Mylar", True, "keyword_rule"
        if "TOUCH" in txt_u or "TP" in re.sub(r"[^A-Z]", " ", txt_u):
            return "Touchpad Mylar", True, "keyword_rule"
        return "Mylar", True, "cn_keyword"

    # 纯英文：大小写统一后去匹配
    norm = smart_title(raw)
    if norm in MYLAR_CATS:
        # 如果只是大小写不同，也算修正
        changed = (norm != raw)
        return norm, changed, "case_fix" if changed else "standard"

    # 模糊匹配到 Mylar 类别（阈值高一些，避免把别的名称误改）
    hit = fuzzy_one(norm, MYLAR_CATS, threshold=86)
    if hit:
        return hit, True, "fuzzy"

    return raw, False, "unchanged"


# 2) 颜色列：只认颜色词，严格模式，避免把名字里“蓝/黄”误改
COLOR_CN = {
    "黑": "Black", "黑色": "Black",
    "黄": "Yellow", "黄色": "Yellow",
    "蓝": "Blue", "蓝色": "Blue"
}
COLOR_STRICT_RE = re.compile(r"^\s*(黑色?|黄色?|蓝色?|black|yellow|blue)\s*$", re.IGNORECASE)


def clean_color(val: str):
    if pd.isna(val) or not str(val).strip():
        return val, False, "none"
    raw = str(val).strip()

    # 严格：整格只含颜色词才替换，避免误清洗
    if COLOR_STRICT_RE.match(raw):
        # 中文直接映射；英文统一大小写到标准
        key = raw.lower()
        if key in {"black", "yellow", "blue"}:
            std = key.capitalize()
            return std, (std != raw), "case_fix" if std != raw else "standard"
        # 中文
        mapped = COLOR_CN.get(raw, None)
        if mapped:
            return mapped, True, "cn_mapping"

    # 尝试与标准颜色模糊匹配（非常高阈值）
    hit = fuzzy_one(raw, COLORS, threshold=92)
    if hit:
        return hit, True, "fuzzy"

    return raw, False, "unchanged"


# 3) 材质列：严格识别常用中文与英文简称
MATERIAL_CN = {
    "聚碳酸酯": "PC",
    "聚酯": "PET",
    "丙烯酸": "Acrylic",
    "聚氯乙烯": "PVC",
}
MATERIAL_STRICT_RE = re.compile(r"^\s*(pc|pet|pvc|acrylic|聚碳酸酯|聚酯|丙烯酸|聚氯乙烯)\s*$", re.IGNORECASE)


def clean_material(val: str):
    if pd.isna(val) or not str(val).strip():
        return val, False, "none"
    raw = str(val).strip()
    if MATERIAL_STRICT_RE.match(raw):
        # 英文
        low = raw.lower()
        if low in {"pc", "pet", "pvc"}:
            std = low.upper()
            return std, (std != raw), "case_fix" if std != raw else "standard"
        if low == "acrylic":
            return "Acrylic", (raw != "Acrylic"), "case_fix" if raw != "Acrylic" else "standard"
        # 中文
        mapped = MATERIAL_CN.get(raw, None)
        if mapped:
            return mapped, True, "cn_mapping"

    hit = fuzzy_one(raw, MATERIALS, threshold=92)
    if hit:
        return hit, True, "fuzzy"

    return raw, False, "unchanged"


# 4) 指纹列：各种写法统一为 W/ FP / W/O FP
def clean_fingerprint(val: str):
    if pd.isna(val) or not str(val).strip():
        return val, False, "none"
    raw = str(val).strip().upper().replace(" ", "")
    # 常见中文
    if raw in {"有", "YES", "Y"}: return "W/ FP", True, "mapping"
    if raw in {"无", "NO", "N"}:  return "W/O FP", True, "mapping"
    # 英文变体
    if "W/FP" in raw or "WITHFP" in raw or "HASFP" in raw:
        return "W/ FP", True, "mapping"
    if "W/OFP" in raw or "WITHOUTFP" in raw or "NOFP" in raw:
        return "W/O FP", True, "mapping"

    # 已经是标准？
    if raw == "W/FP":  return "W/ FP", ("W/ FP" != str(val)), "case_fix" if ("W/ FP" != str(val)) else "standard"
    if raw == "W/OFP": return "W/O FP", ("W/O FP" != str(val)), "case_fix" if ("W/O FP" != str(val)) else "standard"

    return str(val), False, "unchanged"


# 5) 胶型列：严格只认标准清单，模糊阈值很高，避免误改
def clean_adhesive(val: str):
    if pd.isna(val) or not str(val).strip():
        return val, False, "none"
    raw = str(val).strip()
    if raw in ADHESIVES:
        return raw, False, "standard"
    # 去掉可能的空格/大小写
    norm = raw.upper().replace(" ", "")
    for a in ADHESIVES:
        if norm == a.upper().replace(" ", ""):
            return a, (a != raw), "case_fix" if (a != raw) else "standard"
    hit = fuzzy_one(raw, ADHESIVES, threshold=94)  # 非常高，谨慎
    if hit:
        return hit, True, "fuzzy"
    return raw, False, "unchanged"


# 6) 其它特殊属性：目前只把“Gluing/胶合/粘接”等统一为 Gluing，其余不动
def clean_other(val: str):
    if pd.isna(val) or not str(val).strip():
        return val, False, "none"
    raw = str(val).strip()
    u = raw.upper()
    if "GLUING" in u or "GLUE" in u or "ADHESIVE" in u or "BOND" in u:
        return "Gluing", (raw != "Gluing"), "mapping" if (raw != "Gluing") else "standard"
    if "胶合" in raw or "粘接" in raw or "粘合" in raw:
        return "Gluing", True, "cn_mapping"
    if raw in OTHERS:
        return raw, False, "standard"
    return raw, False, "unchanged"


# 7) 三个尺寸列：提取数字，保留 3 位小数；提取失败则不动、不卡
NUM_RE = re.compile(r"-?\d+(?:\.\d+)?")


def clean_number(val: str):
    if pd.isna(val) or str(val).strip() == "":
        return val, False, "none"
    s = str(val)
    m = NUM_RE.search(s.replace(",", ""))
    if not m:
        return s, False, "unchanged"
    num = float(m.group())
    fixed = f"{num:.3f}".rstrip("0").rstrip(".")
    return fixed, (fixed != s), "number_extract" if (fixed != s) else "standard"


# ====== 按列派发器 ======
def clean_cell(col_name: str, val: str):
    if col_name in ("物料简称", "名称"):
        return clean_mylar_name(val)
    if col_name == "颜色":
        return clean_color(val)
    if col_name == "材质":
        return clean_material(val)
    if col_name == "是否带指纹":
        return clean_fingerprint(val)
    if col_name == "背胶型号":
        return clean_adhesive(val)
    if col_name == "其它特殊属性":
        return clean_other(val)
    if col_name in ("长L(mm)", "宽W(mm)", "厚H(mm)"):
        return clean_number(val)
    # 其他列不应进入这里；但为安全返回原值
    return val, False, "none"


# ====== APP 主体（仅在原表上标绿，无新增列行） ======
st.title("📂 指定列清洗（安全模式：避免误清洗）")

uploaded_file = st.file_uploader("上传文件（CSV / XLS / XLSX）", type=["csv", "xls", "xlsx"])
if uploaded_file:
    # 读入
    if uploaded_file.name.lower().endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str)
    else:
        df = pd.read_excel(uploaded_file, dtype=str)

    st.write("原始数据预览：", df.head())

    # 检查需要清洗的列是否存在
    missing = [c for c in TARGET_COLS if c not in df.columns]
    if missing:
        st.warning(f"下列目标列在文件中未找到：{missing}。只会清洗存在的列。")

    if st.button("开始清洗并下载"):
        cell_colors = []  # (row, col)
        changes = []  # 日志

        # 遍历：仅清洗存在于表中的目标列
        for col_name in [c for c in TARGET_COLS if c in df.columns]:
            col_idx = df.columns.get_loc(col_name) + 1  # Excel 列序号（1-based）
            new_col = []
            for row_idx, val in enumerate(df[col_name], start=2):  # Excel 行号从2开始（1是表头）
                new_val, changed, rule = clean_cell(col_name, val)
                new_col.append(new_val)
                if changed:
                    cell_colors.append((row_idx, col_idx))
                    changes.append([col_name, row_idx, val, new_val, rule])
            df[col_name] = new_col

        # 保存为 Excel
        cleaned_file = "MAM_cleaned_selected.xlsx"
        df.to_excel(cleaned_file, index=False)

        # 只标绿色（被修正）
        if cell_colors:
            wb = load_workbook(cleaned_file)
            ws = wb.active
            green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            for r, c in cell_colors:
                ws.cell(r, c).fill = green
            wb.save(cleaned_file)

        # 日志（单独文件，可选下载）
        if changes:
            log_df = pd.DataFrame(changes, columns=["列名", "行号", "原始值", "修改后", "规则"])
            log_file = "MAM_clean_log.xlsx"
            log_df.to_excel(log_file, index=False)
            st.download_button("📑 下载修改日志", open(log_file, "rb"), file_name=log_file)

        st.success("✅ 清洗完成（仅对指定列，且只在原表上标绿）")
        st.download_button("⬇️ 下载清洗后的文件", open(cleaned_file, "rb"), file_name=cleaned_file)
